# output_handler.py
import pandas as pd
from openpyxl import load_workbook
from pandas import ExcelWriter


def save_responses_to_excel(data, filepath, csv_mode=False):
    """
    Save the prompt-response pairs to either CSV or Excel file.
    Tags each response with a source label.

    Parameters:
    - data: list of dicts like
      [{'prompt': str, 'response': str, 'source': 'pubmed'|'clinical_trial'|'other'}, ...]
    - filepath: output file path string
    - csv_mode: if True, save as CSV, else Excel
    """
    # Convert to DataFrame
    df = pd.DataFrame(data)

    # Validate required columns
    required = {'prompt', 'response', 'source'}
    if not required.issubset(df.columns):
        raise ValueError(f"Data dicts must contain keys: {required}")

    # Reorder columns if needed
    df = df[['prompt', 'response', 'source']]

    if csv_mode:
        if not filepath.lower().endswith('.csv'):
            filepath += '.csv'
        df.to_csv(filepath, index=False)
        print(f"\nResponses saved as CSV to {filepath}")
        return

    try:
        book = load_workbook(filepath)
        writer = ExcelWriter(filepath, engine="openpyxl")
        writer.book = book
        startrow = book['Sheet1'].max_row if 'Sheet1' in book.sheetnames else 0
        df.to_excel(writer, sheet_name="Sheet1", index=False, header=startrow == 0, startrow=startrow)
        writer.save()
        writer.close()
        print(f"\nResponses appended to Excel file {filepath}")
    except FileNotFoundError:
        # Create new Excel file if doesn't exist
        df.to_excel(filepath, index=False)
        print(f"\nResponses saved as new Excel file {filepath}")


def save_responses_to_csv(data, filepath):
    """
    Convenience wrapper to save responses as CSV.
    """
    save_responses_to_excel(data, filepath, csv_mode=True)

def convert_to_prompt_response_format(data):
    """
    Converts nested data into a list of prompt-response dicts for Excel/CSV output.
    Assumes new structure:
    {
        "nct_id": "...",
        "sources": {
            "clinical_trials": { "source": ..., "data": {...} },
            "pubmed": { "source": ..., "matches_found": bool, "pmids": [...], "studies": [...] }
        }
    }
    """
    records = []

    ct_data = data.get("sources", {}).get("clinical_trials", {}).get("data", {})
    if ct_data:
        records.append({
            "prompt": "Clinical trial data",
            "response": json.dumps(ct_data, indent=2),
            "source": "clinical_trials"
        })

    pubmed = data.get("sources", {}).get("pubmed", {})
    studies = pubmed.get("studies", [])
    if studies:
        for study in studies:
            title = study.get("title", "PubMed Study")
            records.append({
                "prompt": f"PubMed article: {title}",
                "response": json.dumps(study, indent=2),
                "source": "pubmed"
            })
    else:
        records.append({
            "prompt": "PubMed enrichment",
            "response": "No PubMed matches found.",
            "source": "pubmed"
        })

    return records
