import pandas as pd
from openpyxl import load_workbook
from pandas import ExcelWriter

def save_responses_to_excel(data, filepath, csv_mode=False):
    """
    Save the prompt-response pairs to either CSV or Excel file.

    Parameters:
    - data: list of tuples [(prompt, response), ...]
    - filepath: output file path string
    - csv_mode: if True, save as CSV, else Excel
    """

    df = pd.DataFrame(data, columns=["Prompt", "Response"])

    if csv_mode:
        # Ensure .csv extension
        if not filepath.lower().endswith('.csv'):
            filepath += '.csv'
        df.to_csv(filepath, index=False)
        print(f"\nResponses saved as CSV to {filepath}")
        return

    # Save or append to Excel file
    try:
        book = load_workbook(filepath)
        writer = ExcelWriter(filepath, engine="openpyxl")
        writer.book = book
        # Append to existing sheet 'Sheet1'
        startrow = book['Sheet1'].max_row if 'Sheet1' in book.sheetnames else 0
        df.to_excel(writer, sheet_name="Sheet1", index=False, header=startrow == 0, startrow=startrow)
        writer.save()
        writer.close()
        print(f"\nResponses appended to Excel file {filepath}")
    except FileNotFoundError:
        # File does not exist; create new
        df.to_excel(filepath, index=False)
        print(f"\nResponses saved as new Excel file {filepath}")

def save_responses_to_csv(data, filepath):
    """
    Convenience wrapper to save responses as CSV.
    """
    save_responses_to_excel(data, filepath, csv_mode=True)
