#!/usr/bin/env python3
"""
Concordance Test: Agent (JSON job results) vs Human Annotators (R1 & R2)

Aggregates agent annotations from multiple overnight JSON job files,
matches against two independent human replicates from the clinical
trials Excel file, and computes concordance metrics.

Reports:
  - Agent vs R1, Agent vs R2, R1 vs R2
  - Raw agreement %, Cohen's kappa per field
  - Detailed disagreements for error analysis
"""

import json
import os
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent.parent
EXCEL_PATH = Path(
    "/Users/amphoraxe/Developer/amphoraxe/dev-llm.amphoraxe.ca/docs/"
    "clinical_trials-with-sequences.xlsx"
)
JSON_DIR = BASE_DIR / "results" / "json"
OUTPUT_DIR = BASE_DIR / "results" / "concordance"

# The three overnight job result files (Mar 16)
JOB_FILES = [
    "219670901c18.json",
    "dbc48eb94a68.json",
    "a96defa1fb73.json",
]

# ---------------------------------------------------------------------------
# Field definitions  (Excel column indices are 0-based)
# ---------------------------------------------------------------------------
FIELDS = {
    "classification": {
        "excel_r1_col": 10,  # K
        "excel_r2_col": 10,  # K
        "blank_means_skip": True,
    },
    "delivery_mode": {
        "excel_r1_col": 12,  # M
        "excel_r2_col": 12,  # M
        "blank_means_skip": True,
    },
    "outcome": {
        "excel_r1_col": 17,  # R
        "excel_r2_col": 17,  # R
        "blank_means_skip": True,
    },
    "reason_for_failure": {
        "excel_r1_col": 18,  # S
        "excel_r2_col": 18,  # S
        "blank_means_skip": False,  # blank IS valid
    },
    "peptide": {
        "excel_r1_col": 21,  # V
        "excel_r2_col": 21,  # V
        "blank_means_skip": True,
    },
    "sequence": {
        "excel_r1_col": 13,  # N (Sequence column)
        "excel_r2_col": 13,  # N
        "blank_means_skip": True,
    },
}

# ---------------------------------------------------------------------------
# Value normalisation
# ---------------------------------------------------------------------------
DELIVERY_MODE_ALIASES = {
    "intravenous": "IV",
    "iv": "IV",
    "injection/infusion - intravenous": "IV",
    "oral - unspecified": "Oral - Unspecified",
    "oral - capsule": "Oral - Capsule",
    "oral - tablet": "Oral - Tablet",
    "oral - drink": "Oral - Drink",
    "oral - food": "Oral - Food",
    "injection": "Injection/Infusion - Other/Unspecified",
    "topical - unspecified": "Topical - Unspecified",
    "other/unspecified": "Other/Unspecified",
    "sc": "Injection/Infusion - Subcutaneous/Intradermal",
    "subcutaneous": "Injection/Infusion - Subcutaneous/Intradermal",
    "intradermal": "Injection/Infusion - Subcutaneous/Intradermal",
    "im": "Injection/Infusion - Intramuscular",
    "intramuscular": "Injection/Infusion - Intramuscular",
}

OUTCOME_ALIASES = {
    "active": "Active, not recruiting",
    "active, not recruiting": "Active, not recruiting",
    "active not recruiting": "Active, not recruiting",
    "recruiting": "Recruiting",
    "failed": "Failed - completed trial",
    "failed - completed trial": "Failed - completed trial",
    "completed": "Failed - completed trial",
    "positive": "Positive",
    "terminated": "Terminated",
    "withdrawn": "Withdrawn",
    "unknown": "Unknown",
}

CLASSIFICATION_ALIASES = {
    "amp(infection)": "AMP(infection)",
    "amp(other)": "AMP(other)",
    "amp (infection)": "AMP(infection)",
    "amp (other)": "AMP(other)",
    "other": "Other",
}

REASON_ALIASES = {
    "business reason": "Business Reason",
    "business_reason": "Business Reason",
    "ineffective for purpose": "Ineffective for purpose",
    "ineffective_for_purpose": "Ineffective for purpose",
    "recruitment issues": "Recruitment issues",
    "recruitment_issues": "Recruitment issues",
    "toxic/unsafe": "Toxic/Unsafe",
    "toxic_unsafe": "Toxic/Unsafe",
    "due to covid": "Due to covid",
    "due_to_covid": "Due to covid",
}

PEPTIDE_ALIASES = {
    "true": "True",
    "false": "False",
    "yes": "True",
    "no": "False",
    "1": "True",
    "0": "False",
}


def _normalise_sequence(s: str) -> str:
    """Normalise an amino acid sequence string for concordance comparison.

    Handles human annotation formats: spaced every 5 chars, modification
    markers, D-amino acid prefixes, pipe-separated multi-sequences,
    three-letter code detection.
    """
    import re as _re

    if not s or not s.strip():
        return ""

    # Detect three-letter code sequences (e.g., Cpa-c[d-Cys-Aph(Hor)...])
    # These contain 3-letter AA codes like Cys, Ala, Gly, Lys, Thr, Phe
    _THREE_LETTER_CODES = {
        "ala", "arg", "asn", "asp", "cys", "gln", "glu", "gly", "his",
        "ile", "leu", "lys", "met", "phe", "pro", "ser", "thr", "trp",
        "tyr", "val", "aph", "cbm",
    }
    s_lower = s.lower()
    three_letter_count = sum(1 for code in _THREE_LETTER_CODES if code in s_lower)
    # If contains 3+ three-letter codes AND no long AA stretch, it's three-letter notation
    has_long_aa = bool(_re.search(r"[A-Z]{5,}", s))
    if three_letter_count >= 3 and not has_long_aa:
        return "[three-letter notation]"

    # Handle pipe-separated multi-sequences: normalize each part, sort, rejoin
    if "|" in s:
        parts = [p.strip() for p in s.split("|") if p.strip()]
        normalised_parts = []
        for part in parts:
            norm = _normalise_single_sequence(part)
            if norm and norm != "[three-letter notation]":
                normalised_parts.append(norm)
        normalised_parts.sort()
        return " | ".join(normalised_parts) if normalised_parts else ""

    return _normalise_single_sequence(s)


def _normalise_single_sequence(s: str) -> str:
    """Normalise a single amino acid sequence (no pipe separators)."""
    import re as _re

    if not s or not s.strip():
        return ""

    cleaned = s.strip()

    # Remove modification prefixes: Ac-, H-, Fmoc-, Boc-, cyclo(
    cleaned = _re.sub(r"^(Ac-|H-|Fmoc-|Boc-|cyclo\()", "", cleaned, flags=_re.IGNORECASE)

    # Remove modification suffixes: -NH2, -OH, -COOH, -amide, -acid, (ol)
    # Also handle Unicode: -NH₂
    cleaned = _re.sub(r"(-NH[2₂]|-OH|-COOH|-amide|-acid|\(ol\))$", "", cleaned, flags=_re.IGNORECASE)

    # Remove all whitespace, tabs, newlines
    cleaned = _re.sub(r"\s+", "", cleaned)

    # Strip D-amino acid prefixes: dF → F, dC → C (lowercase d before uppercase AA)
    cleaned = _re.sub(r"(?<![a-zA-Z])d([A-Z])", r"\1", cleaned)

    # Convert single-letter dash notation: K-K-W-W-K → KKWWK
    if _re.match(r"^[A-Z]-[A-Z]-", cleaned):
        cleaned = cleaned.replace("-", "")

    # Uppercase
    cleaned = cleaned.upper()

    # Keep only valid AA characters
    cleaned = _re.sub(r"[^ACDEFGHIKLMNPQRSTVWYBZXUOJ]", "", cleaned)

    return cleaned if len(cleaned) >= 2 else ""


def normalise(value, field_name):
    """Normalise a value for comparison, returning (normalised_str, is_blank)."""
    if value is None:
        return ("", True)

    # Handle booleans (Excel stores Peptide as bool)
    if isinstance(value, bool):
        return (str(value), False)

    s = str(value).strip()
    if s == "" or s.lower() in ("none", "n/a"):
        return ("", True)

    s_lower = s.lower()

    if field_name == "delivery_mode":
        # For multi-value delivery modes, normalise each part
        parts = [p.strip() for p in s.split(",")]
        normalised_parts = []
        for part in parts:
            part_lower = part.strip().lower()
            normalised_parts.append(
                DELIVERY_MODE_ALIASES.get(part_lower, part.strip())
            )
        normalised_parts.sort()
        return (", ".join(normalised_parts), False)
    elif field_name == "outcome":
        return (OUTCOME_ALIASES.get(s_lower, s), False)
    elif field_name == "classification":
        return (CLASSIFICATION_ALIASES.get(s_lower, s), False)
    elif field_name == "reason_for_failure":
        return (REASON_ALIASES.get(s_lower, s), False)
    elif field_name == "peptide":
        return (PEPTIDE_ALIASES.get(s_lower, s), False)
    elif field_name == "sequence":
        return (_normalise_sequence(s), False)
    else:
        return (s, False)


def _bucket_value(normalised_value: str, field_name: str) -> str:
    """Bucket normalised values into broad categories for concordance.

    Conforms to how concordance is measured:
    - Classification: AMP(infection) + AMP(other) → AMP
    - Delivery Mode: injection/infusion, oral, topical, other
    - Outcome: active + recruiting → Active
    - Peptide: blank → True (assume peptide unless explicitly False)
    """
    if not normalised_value:
        if field_name == "peptide":
            return "True"  # Default: assume peptide
        return normalised_value

    s = normalised_value

    if field_name == "classification":
        if s.startswith("AMP"):
            return "AMP"
        return s

    elif field_name == "delivery_mode":
        s_lower = s.lower()
        if "injection" in s_lower or "infusion" in s_lower or s_lower == "iv":
            return "Injection/Infusion"
        if "oral" in s_lower:
            return "Oral"
        if "topical" in s_lower:
            return "Topical"
        if "intranasal" in s_lower:
            return "Intranasal"
        if "inhalation" in s_lower:
            return "Inhalation"
        return "Other"

    elif field_name == "outcome":
        if s in ("Active, not recruiting", "Recruiting"):
            return "Active"
        return s

    elif field_name == "peptide":
        return s

    return s


# ---------------------------------------------------------------------------
# Statistical functions (from shared stats module)
# ---------------------------------------------------------------------------
# Add project root to path so we can import app.services
sys.path.insert(0, str(BASE_DIR))

from app.services.concordance_stats import (
    cohens_kappa,
    kappa_confidence_interval,
    gwets_ac1_with_ci,
    prevalence_index,
    bias_index,
    landis_koch_interpretation as kappa_interpretation_stats,
)


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------
def load_excel_annotations(path):
    """Load human annotations from both replicate sheets.
    Returns dict: { nct_id: { 'r1': {field: raw_value}, 'r2': {field: raw_value} } }
    """
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    data = defaultdict(lambda: {"r1": {}, "r2": {}})

    for replicate, sheet_name in [
        ("r1", "Trials Replicate 1"),
        ("r2", "Trials Replicate 2"),
    ]:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            nct = row[0]
            if nct is None:
                continue
            nct = str(nct).strip()
            for field_name, field_def in FIELDS.items():
                col_idx = field_def[f"excel_{replicate}_col"]
                raw = row[col_idx] if col_idx < len(row) else None
                data[nct][replicate][field_name] = raw

    wb.close()
    return dict(data)


def load_excel_row_to_nct(path):
    """Load mapping of (sheet_name, data_row_number) -> nct_id.

    data_row_number is 1-based (row 1 = first data row after header).
    Returns: { sheet_name: { row_number: nct_id } }
    """
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    mapping = {}

    for sheet_name in ["Trials Replicate 1", "Trials Replicate 2"]:
        ws = wb[sheet_name]
        sheet_map = {}
        data_row = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            data_row += 1
            nct = row[0]
            if nct is None:
                continue
            nct = str(nct).strip()
            sheet_map[data_row] = nct
        mapping[sheet_name] = sheet_map

    wb.close()
    return mapping


def load_agent_annotations_from_json(json_paths):
    """Load agent annotations from one or more JSON job result files.

    Uses the verification.fields[].final_value for each annotation field.
    When a trial appears in multiple jobs, the LAST occurrence wins
    (later jobs are assumed more recent / refined).

    Returns dict: { nct_id: {field: raw_value} }
    """
    data = {}
    total_trials = 0
    duplicates = 0

    for path in json_paths:
        with open(str(path), "r") as f:
            job = json.load(f)

        trials = job.get("trials", [])
        job_name = Path(path).stem
        print(f"  {job_name}: {len(trials)} trials")

        for trial in trials:
            nct_id = trial.get("nct_id", "")
            if not nct_id:
                continue

            total_trials += 1
            if nct_id in data:
                duplicates += 1

            # Extract final_value from verification.fields
            entry = {}
            verification = trial.get("verification", {})
            fields_list = verification.get("fields", [])

            for field_obj in fields_list:
                fname = field_obj.get("field_name", "")
                fval = field_obj.get("final_value", "")
                if fname in FIELDS:
                    entry[fname] = fval

            # Also try annotations array as fallback for missing fields
            for anno in trial.get("annotations", []):
                fname = anno.get("field_name", "")
                if fname in FIELDS and fname not in entry:
                    entry[fname] = anno.get("value", "")

            # Fill missing fields with empty
            for fname in FIELDS:
                if fname not in entry:
                    entry[fname] = ""

            data[nct_id] = entry

    print(f"  Total trial records: {total_trials}")
    print(f"  Duplicates (last wins): {duplicates}")
    print(f"  Unique NCT IDs: {len(data)}")

    return data


# ---------------------------------------------------------------------------
# Concordance computation
# ---------------------------------------------------------------------------
def compute_concordance(agent_data, human_data):
    """
    For each field, compute concordance between:
      - Agent vs R1
      - Agent vs R2
      - R1 vs R2

    Returns (results_dict, all_disagreements, common_ncts).
    """
    agent_ncts = set(agent_data.keys())
    human_ncts = set(human_data.keys())
    common_ncts = sorted(agent_ncts & human_ncts)

    print(f"Agent NCT IDs:  {len(agent_ncts)}")
    print(f"Human NCT IDs:  {len(human_ncts)}")
    print(f"Overlapping:    {len(common_ncts)}")

    # Show agent-only and human-only
    agent_only = sorted(agent_ncts - human_ncts)
    human_only_sample = sorted(human_ncts - agent_ncts)[:5]
    if agent_only:
        print(f"Agent-only ({len(agent_only)}): {agent_only[:10]}{'...' if len(agent_only) > 10 else ''}")
    if human_only_sample:
        print(f"Human-only ({len(human_ncts - agent_ncts)}): {human_only_sample}...")
    print()

    if not common_ncts:
        print("ERROR: No overlapping NCT IDs found!")
        return {}, [], []

    results = {}
    all_disagreements = []

    for field_name, field_def in FIELDS.items():
        blank_means_skip = field_def["blank_means_skip"]
        field_result = {}

        for pair_name, src_a_label, src_b_label, get_a, get_b in [
            (
                "Agent vs R1",
                "Agent",
                "R1",
                lambda nct: agent_data[nct].get(field_name, ""),
                lambda nct: human_data[nct]["r1"].get(field_name),
            ),
            (
                "Agent vs R2",
                "Agent",
                "R2",
                lambda nct: agent_data[nct].get(field_name, ""),
                lambda nct: human_data[nct]["r2"].get(field_name),
            ),
            (
                "R1 vs R2",
                "R1",
                "R2",
                lambda nct: human_data[nct]["r1"].get(field_name),
                lambda nct: human_data[nct]["r2"].get(field_name),
            ),
        ]:
            labels_a = []
            labels_b = []
            pair_disagreements = []
            skipped_blank = 0

            for nct in common_ncts:
                raw_a = get_a(nct)
                raw_b = get_b(nct)

                norm_a, blank_a = normalise(raw_a, field_name)
                norm_b, blank_b = normalise(raw_b, field_name)

                # Skip if blank_means_skip and EITHER side is blank
                if blank_means_skip and (blank_a or blank_b):
                    skipped_blank += 1
                    continue

                # For reason_for_failure: blank is valid
                if not blank_means_skip:
                    if blank_a:
                        norm_a = ""
                    if blank_b:
                        norm_b = ""

                # v14: For sequence field, use tiered matching
                # (exact, substring, no match). For concordance stats,
                # treat substring matches as agreement.
                if field_name == "sequence" and norm_a and norm_b and norm_a != norm_b:
                    # Check substring relationship (one contains the other)
                    if norm_a in norm_b or norm_b in norm_a:
                        # Treat as match for concordance, but track separately
                        labels_a.append(norm_a)
                        labels_b.append(norm_a)  # Force agreement for stats
                        pair_disagreements.append(
                            {
                                "nct_id": nct,
                                "field": field_name,
                                "comparison": pair_name,
                                "match_type": "substring",
                                f"{src_a_label.lower()}_value": norm_a,
                                f"{src_b_label.lower()}_value": norm_b,
                                f"{src_a_label.lower()}_raw": str(raw_a),
                                f"{src_b_label.lower()}_raw": str(raw_b),
                            }
                        )
                        continue

                labels_a.append(norm_a)
                labels_b.append(norm_b)

                if norm_a != norm_b:
                    pair_disagreements.append(
                        {
                            "nct_id": nct,
                            "field": field_name,
                            "comparison": pair_name,
                            f"{src_a_label.lower()}_value": norm_a,
                            f"{src_b_label.lower()}_value": norm_b,
                            f"{src_a_label.lower()}_raw": str(raw_a),
                            f"{src_b_label.lower()}_raw": str(raw_b),
                        }
                    )

            n = len(labels_a)
            if n > 0:
                kappa, po, pe = cohens_kappa(labels_a, labels_b)
                kappa_k, kappa_ci_lo, kappa_ci_hi = kappa_confidence_interval(labels_a, labels_b)
                ac1, ac1_ci_lo, ac1_ci_hi = gwets_ac1_with_ci(labels_a, labels_b)
                pi = prevalence_index(labels_a, labels_b)
                bi = bias_index(labels_a, labels_b)
                agreements = sum(
                    1 for a, b in zip(labels_a, labels_b) if a == b
                )
            else:
                kappa, po, pe = float("nan"), float("nan"), float("nan")
                kappa_ci_lo, kappa_ci_hi = float("nan"), float("nan")
                ac1, ac1_ci_lo, ac1_ci_hi = float("nan"), float("nan"), float("nan")
                pi, bi = float("nan"), float("nan")
                agreements = 0

            field_result[pair_name] = {
                "n": n,
                "skipped_blank": skipped_blank,
                "agreements": agreements,
                "raw_agreement_pct": round(po * 100, 1) if n > 0 else None,
                "cohens_kappa": round(kappa, 4) if n > 0 else None,
                "kappa_ci": (kappa_ci_lo, kappa_ci_hi) if n > 0 else None,
                "ac1": ac1 if n > 0 else None,
                "ac1_ci": (ac1_ci_lo, ac1_ci_hi) if n > 0 else None,
                "prevalence_index": pi if n > 0 else None,
                "bias_index": bi if n > 0 else None,
                "pe": round(pe, 4) if n > 0 else None,
                "disagreements": pair_disagreements,
            }
            all_disagreements.extend(pair_disagreements)

        results[field_name] = field_result

    return results, all_disagreements, common_ncts


# ---------------------------------------------------------------------------
# Concordance v3: Three-tier analysis
# ---------------------------------------------------------------------------
FAILURE_OUTCOMES = {"Terminated", "Withdrawn", "Failed - completed trial"}
NON_FAILURE_OUTCOMES = {"Positive", "Recruiting", "Active, not recruiting", "Unknown"}


def _is_reason_blank_v3(reason_raw, outcome_raw, field_name):
    """Outcome-aware blank handling for reason_for_failure (v3).

    Returns (normalised_value, treat_as_blank, skip_entirely).
    - skip_entirely: both outcome and reason are blank → annotator didn't engage
    - treat_as_blank: failure outcome + blank reason → missing data
    - Otherwise: return normalised value (may be empty string for non-failure outcomes)
    """
    norm_reason, reason_blank = normalise(reason_raw, field_name)
    norm_outcome, outcome_blank = normalise(outcome_raw, "outcome")

    if outcome_blank and reason_blank:
        # Annotator didn't engage with this trial at all
        return ("", True, True)

    if not reason_blank:
        # Reason is filled — always use it
        return (norm_reason, False, False)

    # Reason is blank — check outcome to decide if legitimate
    if norm_outcome in NON_FAILURE_OUTCOMES:
        # Non-failure outcome + blank reason → legitimate "no failure"
        return ("", False, False)

    if norm_outcome in FAILURE_OUTCOMES:
        # Failure outcome + blank reason → missing data
        return ("", True, False)

    # Fallback: unknown outcome value — treat as blank
    return ("", True, False)


def compute_concordance_v3(agent_data, human_data, common_ncts):
    """
    Three-tier concordance analysis for all fields.

    Tier 1 (Strict): Both sides filled. Same as v2.
    Tier 2 (Coverage-adjusted): At least one side filled. One-sided blank = disagree.
    Tier 3 (Full population): All overlapping trials. Both-blank = agree.

    Returns (v3_results, coverage_report).
    """
    v3_results = {}
    coverage_report = {}

    for field_name, field_def in FIELDS.items():
        field_result = {}
        is_reason = (field_name == "reason_for_failure")

        # Build per-source data with v3 blank handling
        # For each NCT, determine: agent value, r1 value, r2 value + blank status
        trial_data = []
        for nct in common_ncts:
            agent_raw = agent_data[nct].get(field_name, "")
            r1_raw = human_data[nct]["r1"].get(field_name)
            r2_raw = human_data[nct]["r2"].get(field_name)

            if is_reason:
                # Outcome-aware blank handling for reason_for_failure
                agent_outcome = agent_data[nct].get("outcome", "")
                r1_outcome = human_data[nct]["r1"].get("outcome")
                r2_outcome = human_data[nct]["r2"].get("outcome")

                a_norm, a_blank, a_skip = _is_reason_blank_v3(agent_raw, agent_outcome, field_name)
                r1_norm, r1_blank, r1_skip = _is_reason_blank_v3(r1_raw, r1_outcome, field_name)
                r2_norm, r2_blank, r2_skip = _is_reason_blank_v3(r2_raw, r2_outcome, field_name)
            else:
                a_norm, a_blank = normalise(agent_raw, field_name)
                r1_norm, r1_blank = normalise(r1_raw, field_name)
                r2_norm, r2_blank = normalise(r2_raw, field_name)
                a_skip = a_blank and field_def["blank_means_skip"]
                r1_skip = r1_blank and field_def["blank_means_skip"]
                r2_skip = r2_blank and field_def["blank_means_skip"]

            trial_data.append({
                "nct": nct,
                "agent": (a_norm, a_blank, a_skip),
                "r1": (r1_norm, r1_blank, r1_skip),
                "r2": (r2_norm, r2_blank, r2_skip),
            })

        # Coverage stats (for R1 and R2 only, since agent always fills)
        r1_filled = sum(1 for t in trial_data if not t["r1"][1])
        r2_filled = sum(1 for t in trial_data if not t["r2"][1])
        both_filled_r1r2 = sum(1 for t in trial_data if not t["r1"][1] and not t["r2"][1])
        both_blank_r1r2 = sum(1 for t in trial_data if t["r1"][1] and t["r2"][1])
        r1_only = sum(1 for t in trial_data if not t["r1"][1] and t["r2"][1])
        r2_only = sum(1 for t in trial_data if t["r1"][1] and not t["r2"][1])

        coverage_report[field_name] = {
            "total": len(common_ncts),
            "r1_filled": r1_filled,
            "r2_filled": r2_filled,
            "both_filled": both_filled_r1r2,
            "both_blank": both_blank_r1r2,
            "r1_only": r1_only,
            "r2_only": r2_only,
        }

        for pair_name, key_a, key_b in [
            ("Agent vs R1", "agent", "r1"),
            ("Agent vs R2", "agent", "r2"),
            ("R1 vs R2", "r1", "r2"),
        ]:
            # Categorise each trial
            both_filled_trials = []
            both_blank_trials = []
            a_only_trials = []
            b_only_trials = []
            both_skip_trials = []

            for t in trial_data:
                val_a, blank_a, skip_a = t[key_a]
                val_b, blank_b, skip_b = t[key_b]

                if skip_a and skip_b:
                    both_skip_trials.append(t)
                elif skip_a and not blank_b:
                    b_only_trials.append(t)
                elif skip_b and not blank_a:
                    a_only_trials.append(t)
                elif not blank_a and not blank_b:
                    both_filled_trials.append(t)
                elif blank_a and blank_b:
                    both_blank_trials.append(t)
                elif not blank_a and blank_b:
                    a_only_trials.append(t)
                elif blank_a and not blank_b:
                    b_only_trials.append(t)
                else:
                    both_skip_trials.append(t)

            # Count agreements among both-filled
            bf_agree = sum(
                1 for t in both_filled_trials
                if t[key_a][0] == t[key_b][0]
            )
            bf_total = len(both_filled_trials)

            # Tier 1: Strict (both filled only)
            t1_n = bf_total
            t1_agree = bf_agree
            t1_pct = round(t1_agree / t1_n * 100, 1) if t1_n > 0 else None

            # Tier 2: Coverage-adjusted (at least one filled)
            t2_n = bf_total + len(a_only_trials) + len(b_only_trials)
            t2_agree = bf_agree  # one-sided blanks count as disagree
            t2_pct = round(t2_agree / t2_n * 100, 1) if t2_n > 0 else None

            # Tier 3: Full population (all overlapping)
            bb_agree = len(both_blank_trials)  # both blank = agreement
            t3_n = bf_total + len(both_blank_trials) + len(a_only_trials) + len(b_only_trials)
            t3_agree = bf_agree + bb_agree
            t3_pct = round(t3_agree / t3_n * 100, 1) if t3_n > 0 else None

            field_result[pair_name] = {
                "tier1": {"n": t1_n, "agreements": t1_agree, "pct": t1_pct},
                "tier2": {"n": t2_n, "agreements": t2_agree, "pct": t2_pct},
                "tier3": {"n": t3_n, "agreements": t3_agree, "pct": t3_pct},
                "counts": {
                    "both_filled": bf_total,
                    "both_blank": len(both_blank_trials),
                    "a_only": len(a_only_trials),
                    "b_only": len(b_only_trials),
                    "both_skip": len(both_skip_trials),
                },
            }

        v3_results[field_name] = field_result

    return v3_results, coverage_report


def compute_concordance_bucketed(agent_data, human_data, common_ncts):
    """Bucketed concordance: broad-category comparison for all fields.

    Uses _bucket_value to collapse fine-grained values into broad categories:
    - Classification: AMP(infection)+AMP(other) → AMP
    - Delivery Mode: 18 values → injection/infusion, oral, topical, intranasal, inhalation, other
    - Outcome: active+recruiting → Active
    - Peptide: blank → True

    Returns dict: {field_name: {pair_name: {n, agreements, pct, kappa}}}
    """
    bucketed_fields = {"classification", "delivery_mode", "outcome", "peptide"}
    results = {}

    for field_name in bucketed_fields:
        field_def = FIELDS[field_name]
        field_result = {}

        for pair_name, get_a, get_b in [
            (
                "Agent vs R1",
                lambda nct: agent_data[nct].get(field_name, ""),
                lambda nct: human_data[nct]["r1"].get(field_name),
            ),
            (
                "Agent vs R2",
                lambda nct: agent_data[nct].get(field_name, ""),
                lambda nct: human_data[nct]["r2"].get(field_name),
            ),
            (
                "R1 vs R2",
                lambda nct: human_data[nct]["r1"].get(field_name),
                lambda nct: human_data[nct]["r2"].get(field_name),
            ),
        ]:
            labels_a = []
            labels_b = []

            for nct in common_ncts:
                raw_a = get_a(nct)
                raw_b = get_b(nct)

                norm_a, blank_a = normalise(raw_a, field_name)
                norm_b, blank_b = normalise(raw_b, field_name)

                # Apply bucketing (handles blanks for peptide)
                buck_a = _bucket_value(norm_a, field_name)
                buck_b = _bucket_value(norm_b, field_name)

                # Skip if either is still blank after bucketing
                if not buck_a or not buck_b:
                    continue

                labels_a.append(buck_a)
                labels_b.append(buck_b)

            n = len(labels_a)
            if n > 0:
                kappa, po, pe = cohens_kappa(labels_a, labels_b)
                agreements = sum(1 for a, b in zip(labels_a, labels_b) if a == b)
            else:
                kappa, po = float("nan"), float("nan")
                agreements = 0

            field_result[pair_name] = {
                "n": n,
                "agreements": agreements,
                "pct": round(po * 100, 1) if n > 0 else None,
                "kappa": round(kappa, 4) if n > 0 else None,
            }

        results[field_name] = field_result

    return results


def print_concordance_bucketed_table(bucketed_results):
    """Print the bucketed concordance summary table."""
    print()
    print("=" * 100)
    print("CONCORDANCE SUMMARY (Bucketed — Broad Categories)")
    print("=" * 100)
    print(
        f"{'Field':<22} {'Comparison':<15} "
        f"{'Agreement':>10} {'Kappa':>8} {'N':>6}"
    )
    print("-" * 100)

    for field_name in ["classification", "delivery_mode", "outcome", "peptide"]:
        field_result = bucketed_results.get(field_name, {})
        first = True
        for pair_name in ["Agent vs R1", "Agent vs R2", "R1 vs R2"]:
            pr = field_result.get(pair_name, {})
            pct_str = f"{pr['pct']:.1f}%" if pr.get("pct") is not None else "N/A"
            kappa_str = f"{pr['kappa']:.4f}" if pr.get("kappa") is not None else "N/A"

            label = field_name if first else ""
            print(
                f"{label:<22} {pair_name:<15} "
                f"{pct_str:>10} {kappa_str:>8} {pr.get('n', 0):>6}"
            )
            first = False
        print("-" * 100)

    print("Buckets: Classification(AMP/Other), Delivery(Inj/Oral/Topical/Intranasal/Inhal/Other), "
          "Outcome(Active=recruiting+active), Peptide(blank=True)")
    print("=" * 100)


def print_concordance_v3_table(v3_results):
    """Print the three-tier concordance summary table."""
    print("=" * 120)
    print("CONCORDANCE SUMMARY (Three-Tier Analysis)")
    print("=" * 120)
    print(
        f"{'Field':<22} {'Comparison':<15} "
        f"{'Tier1(strict)':>14} {'Tier2(coverage)':>16} {'Tier3(full)':>12} "
        f"{'N_strict':>9} {'N_coverage':>11} {'N_full':>7}"
    )
    print("-" * 120)

    for field_name in FIELDS:
        field_result = v3_results.get(field_name, {})
        first = True
        for pair_name in ["Agent vs R1", "Agent vs R2", "R1 vs R2"]:
            pr = field_result.get(pair_name, {})
            t1 = pr.get("tier1", {})
            t2 = pr.get("tier2", {})
            t3 = pr.get("tier3", {})

            t1_str = f"{t1['pct']:.1f}%" if t1.get("pct") is not None else "N/A"
            t2_str = f"{t2['pct']:.1f}%" if t2.get("pct") is not None else "N/A"
            t3_str = f"{t3['pct']:.1f}%" if t3.get("pct") is not None else "N/A"

            label = field_name if first else ""
            print(
                f"{label:<22} {pair_name:<15} "
                f"{t1_str:>14} {t2_str:>16} {t3_str:>12} "
                f"{t1.get('n', 0):>9} {t2.get('n', 0):>11} {t3.get('n', 0):>7}"
            )
            first = False
        print("-" * 120)

    print("=" * 120)


def print_coverage_report(coverage_report, n_total):
    """Print coverage report showing annotation completeness."""
    print(f"\nCOVERAGE REPORT ({n_total}-trial batch)")
    print("=" * 100)
    print(
        f"{'Field':<22} {'R1 filled':>10} {'R2 filled':>10} "
        f"{'Both filled':>12} {'Both blank':>11} {'R1-only':>8} {'R2-only':>8}"
    )
    print("-" * 100)

    for field_name in FIELDS:
        cr = coverage_report.get(field_name, {})
        total = cr.get("total", n_total)
        r1f = cr.get("r1_filled", 0)
        r2f = cr.get("r2_filled", 0)
        bf = cr.get("both_filled", 0)
        bb = cr.get("both_blank", 0)
        r1o = cr.get("r1_only", 0)
        r2o = cr.get("r2_only", 0)
        print(
            f"{field_name:<22} {r1f:>4}/{total:<5} {r2f:>4}/{total:<5} "
            f"{bf:>5}/{total:<5}  {bb:>5}/{total:<5} "
            f"{r1o:>3}/{total:<4} {r2o:>3}/{total:<4}"
        )

    print("=" * 100)


# ---------------------------------------------------------------------------
# Display
# ---------------------------------------------------------------------------
def kappa_interpretation(k):
    """Landis & Koch interpretation of kappa (delegates to stats module)."""
    if k is None:
        return "N/A"
    return kappa_interpretation_stats(k)


def print_concordance_table(results):
    """Print a formatted concordance summary table with CI, AC1, and interpretation."""
    print("=" * 160)
    print(
        f"{'Field':<22} {'Comparison':<15} {'N':>5} {'Skip':>5} "
        f"{'Agree':>6} {'Agree%':>8} {'Kappa':>8} {'95% CI':>16} "
        f"{'AC1':>8} {'Interpretation':<16}"
    )
    print("-" * 160)

    for field_name in FIELDS:
        field_result = results.get(field_name, {})
        first = True
        for pair_name in ["Agent vs R1", "Agent vs R2", "R1 vs R2"]:
            pr = field_result.get(pair_name, {})
            n = pr.get("n", 0)
            skipped = pr.get("skipped_blank", 0)
            agreements = pr.get("agreements", 0)
            pct = pr.get("raw_agreement_pct")
            kappa = pr.get("cohens_kappa")
            kappa_ci = pr.get("kappa_ci")
            ac1 = pr.get("ac1")

            pct_str = f"{pct:.1f}%" if pct is not None else "N/A"
            kappa_str = f"{kappa:.4f}" if kappa is not None else "N/A"
            if kappa_ci is not None:
                ci_str = f"[{kappa_ci[0]:.4f}, {kappa_ci[1]:.4f}]"
            else:
                ci_str = "N/A"
            ac1_str = f"{ac1:.4f}" if ac1 is not None else "N/A"
            interp = kappa_interpretation(kappa)

            label = field_name if first else ""
            print(
                f"{label:<22} {pair_name:<15} {n:>5} {skipped:>5} "
                f"{agreements:>6} {pct_str:>8} {kappa_str:>8} {ci_str:>16} "
                f"{ac1_str:>8} {interp:<16}"
            )
            first = False
        print("-" * 160)

    print("=" * 160)


def print_disagreements(all_disagreements):
    """Print detailed disagreement listing grouped by field and pair."""
    if not all_disagreements:
        print("\nNo disagreements found.")
        return

    print(f"\n{'=' * 115}")
    print(f"DETAILED DISAGREEMENTS ({len(all_disagreements)} total)")
    print(f"{'=' * 115}")

    by_field = defaultdict(lambda: defaultdict(list))
    for d in all_disagreements:
        by_field[d["field"]][d["comparison"]].append(d)

    for field_name in FIELDS:
        field_disags = by_field.get(field_name, {})
        if not field_disags:
            continue
        total_field = sum(len(v) for v in field_disags.values())
        print(f"\n--- {field_name} ({total_field} disagreements) ---")
        for pair_name in ["Agent vs R1", "Agent vs R2", "R1 vs R2"]:
            disags = field_disags.get(pair_name, [])
            if not disags:
                continue
            print(f"  [{pair_name}] ({len(disags)} disagreements)")
            for d in disags:
                val_keys = sorted([k for k in d if k.endswith("_value")])
                val_strs = " | ".join(
                    f"{k.replace('_value', '').upper()}={repr(d[k])}"
                    for k in val_keys
                )
                print(f"    {d['nct_id']}: {val_strs}")


def print_per_trial_matrix(results, common_ncts, agent_data, human_data):
    """Print a per-trial concordance matrix."""
    print(f"\n{'=' * 115}")
    print("PER-TRIAL CONCORDANCE MATRIX")
    print(f"{'=' * 115}")
    print(f"{'NCT ID':<16}", end="")
    for field_name in FIELDS:
        print(f" {'AgR1':>5} {'AgR2':>5} {'R1R2':>5} |", end="")
    print()
    print(f"{'':16}", end="")
    for field_name in FIELDS:
        abbr = field_name[:12]
        width = 20
        print(f" {abbr:^{width}}|", end="")
    print()
    print("-" * (16 + len(FIELDS) * 21))

    for nct in common_ncts:
        print(f"{nct:<16}", end="")
        for field_name, field_def in FIELDS.items():
            blank_means_skip = field_def["blank_means_skip"]

            agent_raw = agent_data[nct].get(field_name, "")
            r1_raw = human_data[nct]["r1"].get(field_name)
            r2_raw = human_data[nct]["r2"].get(field_name)

            a_norm, a_blank = normalise(agent_raw, field_name)
            r1_norm, r1_blank = normalise(r1_raw, field_name)
            r2_norm, r2_blank = normalise(r2_raw, field_name)

            if not blank_means_skip:
                if a_blank:
                    a_norm = ""
                if r1_blank:
                    r1_norm = ""
                if r2_blank:
                    r2_norm = ""

            def match_str(v1, b1, v2, b2):
                if blank_means_skip and (b1 or b2):
                    return "  -  "
                return " YES " if v1 == v2 else " NO  "

            ag_r1 = match_str(a_norm, a_blank, r1_norm, r1_blank)
            ag_r2 = match_str(a_norm, a_blank, r2_norm, r2_blank)
            r1_r2 = match_str(r1_norm, r1_blank, r2_norm, r2_blank)
            print(f" {ag_r1} {ag_r2} {r1_r2} |", end="")
        print()


def print_value_distribution(agent_data, human_data, common_ncts):
    """Print value distribution per field to spot normalisation issues."""
    print(f"\n{'=' * 115}")
    print("VALUE DISTRIBUTIONS (overlapping trials only)")
    print(f"{'=' * 115}")

    for field_name in FIELDS:
        print(f"\n--- {field_name} ---")
        agent_vals = Counter()
        r1_vals = Counter()
        r2_vals = Counter()

        for nct in common_ncts:
            a_norm, a_blank = normalise(agent_data[nct].get(field_name, ""), field_name)
            r1_norm, r1_blank = normalise(human_data[nct]["r1"].get(field_name), field_name)
            r2_norm, r2_blank = normalise(human_data[nct]["r2"].get(field_name), field_name)

            if not a_blank:
                agent_vals[a_norm] += 1
            else:
                agent_vals["<blank>"] += 1
            if not r1_blank:
                r1_vals[r1_norm] += 1
            else:
                r1_vals["<blank>"] += 1
            if not r2_blank:
                r2_vals[r2_norm] += 1
            else:
                r2_vals["<blank>"] += 1

        all_vals = sorted(set(agent_vals) | set(r1_vals) | set(r2_vals))
        print(f"  {'Value':<45} {'Agent':>6} {'R1':>6} {'R2':>6}")
        for v in all_vals:
            print(f"  {v:<45} {agent_vals.get(v, 0):>6} {r1_vals.get(v, 0):>6} {r2_vals.get(v, 0):>6}")


# ---------------------------------------------------------------------------
# Annotator workload attribution
# ---------------------------------------------------------------------------
# Row ranges are 1-based data rows (after header). Ranges are [start, end).
WORKLOAD = {
    "Trials Replicate 1": [
        (1, 309, "Mercan"), (310, 617, "Maya"), (617, 822, "Anat"),
        (823, 926, "Ali"), (926, 1186, "Emre"), (1187, 1417, "Iris"),
        (1417, 1544, "Ali"), (1545, 1846, "Berke"),
    ],
    "Trials Replicate 2": [
        (1, 461, "Emily"), (462, 480, "Anat"), (481, 922, "Emily"),
        (923, 941, "Ali"), (941, 1383, "Emily"), (1384, 1405, "Iris"),
    ],
}


def load_annotator_workload(row_to_nct):
    """Map each (sheet_name, nct_id) to an annotator name based on row ranges.

    Returns: { sheet_name: { nct_id: annotator_name } }
    """
    result = {}
    for sheet_name, ranges in WORKLOAD.items():
        sheet_map = row_to_nct.get(sheet_name, {})
        nct_to_annotator = {}
        for start, end, annotator in ranges:
            for row_num in range(start, end + 1):
                nct = sheet_map.get(row_num)
                if nct:
                    nct_to_annotator[nct] = annotator
        result[sheet_name] = nct_to_annotator
    return result


def get_annotator_for_nct(annotator_map, nct, replicate):
    """Get the annotator name for a given NCT and replicate.

    replicate: 'r1' or 'r2'
    """
    sheet_name = "Trials Replicate 1" if replicate == "r1" else "Trials Replicate 2"
    return annotator_map.get(sheet_name, {}).get(nct, "Unknown")


def print_annotator_analysis(agent_data, human_data, common_ncts, annotator_map):
    """Print per-annotator concordance statistics."""
    import math

    print(f"\n{'=' * 160}")
    print("PER-ANNOTATOR CONCORDANCE ANALYSIS")
    print(f"{'=' * 160}")

    # Build annotator -> list of NCTs for each replicate
    r1_by_annotator = defaultdict(list)
    r2_by_annotator = defaultdict(list)

    for nct in common_ncts:
        r1_ann = get_annotator_for_nct(annotator_map, nct, "r1")
        r2_ann = get_annotator_for_nct(annotator_map, nct, "r2")
        r1_by_annotator[r1_ann].append(nct)
        r2_by_annotator[r2_ann].append(nct)

    # Summary: how many NCTs per annotator
    print("\nAnnotator trial counts (overlapping with agent):")
    print(f"  {'Annotator':<12} {'R1 trials':>10} {'R2 trials':>10} {'Total':>10}")
    print(f"  {'-'*44}")
    all_annotators = sorted(set(list(r1_by_annotator.keys()) + list(r2_by_annotator.keys())))
    for ann in all_annotators:
        r1_n = len(r1_by_annotator.get(ann, []))
        r2_n = len(r2_by_annotator.get(ann, []))
        print(f"  {ann:<12} {r1_n:>10} {r2_n:>10} {r1_n + r2_n:>10}")

    # Per-annotator kappa (Agent vs annotator's subset)
    print(f"\n{'Annotator':<12} {'Rep':<4} {'N_trials':>9} {'Field':<22} "
          f"{'Kappa':>8} {'95% CI':>16} {'AC1':>8} {'Interpretation':<16}")
    print("-" * 110)

    for ann in all_annotators:
        for rep_label, rep_key, nct_map in [
            ("R1", "r1", r1_by_annotator),
            ("R2", "r2", r2_by_annotator),
        ]:
            ncts = nct_map.get(ann, [])
            if not ncts:
                continue

            first_ann = True
            for field_name, field_def in FIELDS.items():
                blank_means_skip = field_def["blank_means_skip"]
                labels_a = []
                labels_b = []

                for nct in ncts:
                    raw_agent = agent_data[nct].get(field_name, "")
                    raw_human = human_data[nct][rep_key].get(field_name)

                    norm_agent, blank_agent = normalise(raw_agent, field_name)
                    norm_human, blank_human = normalise(raw_human, field_name)

                    if blank_means_skip and (blank_agent or blank_human):
                        continue

                    if not blank_means_skip:
                        if blank_agent:
                            norm_agent = ""
                        if blank_human:
                            norm_human = ""

                    labels_a.append(norm_agent)
                    labels_b.append(norm_human)

                n = len(labels_a)
                if n > 0:
                    k, _, _ = cohens_kappa(labels_a, labels_b)
                    _, ci_lo, ci_hi = kappa_confidence_interval(labels_a, labels_b)
                    ac1_val, _, _ = gwets_ac1_with_ci(labels_a, labels_b)
                    interp = kappa_interpretation(k if not math.isnan(k) else None)
                    k_str = f"{k:.4f}" if not math.isnan(k) else "N/A"
                    ci_str = f"[{ci_lo:.4f}, {ci_hi:.4f}]" if not math.isnan(ci_lo) else "N/A"
                    ac1_str = f"{ac1_val:.4f}" if not math.isnan(ac1_val) else "N/A"
                else:
                    k_str = "N/A"
                    ci_str = "N/A"
                    ac1_str = "N/A"
                    interp = "N/A"

                ann_label = ann if first_ann else ""
                rep_show = rep_label if first_ann else ""
                n_show = str(len(ncts)) if first_ann else ""
                first_ann = False

                print(
                    f"{ann_label:<12} {rep_show:<4} {n_show:>9} {field_name:<22} "
                    f"{k_str:>8} {ci_str:>16} {ac1_str:>8} {interp:<16}"
                )
            print("-" * 110)

    print("=" * 160)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print("=" * 115)
    print("AGENT ANNOTATE - CONCORDANCE TEST (JSON OVERNIGHT JOBS)")
    print(f"Timestamp: {datetime.now().isoformat()}")
    print(f"Excel:     {EXCEL_PATH}")
    print(f"JSON dir:  {JSON_DIR}")
    print(f"Job files: {', '.join(JOB_FILES)}")
    print("=" * 115)
    print()

    # Verify files exist
    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found: {EXCEL_PATH}")
        sys.exit(1)

    json_paths = []
    for fname in JOB_FILES:
        p = JSON_DIR / fname
        if not p.exists():
            print(f"ERROR: JSON file not found: {p}")
            sys.exit(1)
        json_paths.append(p)

    # Load data
    print("Loading human annotations from Excel...")
    human_data = load_excel_annotations(EXCEL_PATH)
    print(f"  Loaded {len(human_data)} NCT IDs from Excel")
    print()

    print("Loading agent annotations from JSON job files...")
    agent_data = load_agent_annotations_from_json(json_paths)
    print()

    # Compute concordance
    results, all_disagreements, common_ncts = compute_concordance(
        agent_data, human_data
    )

    if not results:
        sys.exit(1)

    # Print summary table (v2 — kept for backward compatibility)
    print("\nCONCORDANCE SUMMARY (v2 — Strict Only)")
    print_concordance_table(results)

    # Kappa guide
    print("\nKappa Interpretation (Landis & Koch):")
    print("  < 0.00  Poor | 0.00-0.20  Slight | 0.21-0.40  Fair")
    print("  0.41-0.60  Moderate | 0.61-0.80  Substantial | 0.81-1.00  Almost Perfect")
    print()

    # v3 three-tier analysis
    v3_results, coverage_report = compute_concordance_v3(
        agent_data, human_data, common_ncts
    )
    print()
    print_concordance_v3_table(v3_results)
    print_coverage_report(coverage_report, len(common_ncts))

    # v15 bucketed concordance (broad categories)
    bucketed_results = compute_concordance_bucketed(
        agent_data, human_data, common_ncts
    )
    print_concordance_bucketed_table(bucketed_results)

    # Value distributions
    print_value_distribution(agent_data, human_data, common_ncts)

    # Per-trial matrix
    print_per_trial_matrix(results, common_ncts, agent_data, human_data)

    # Disagreements
    print_disagreements(all_disagreements)

    # Per-annotator analysis
    print("\nLoading row-to-NCT mapping for annotator attribution...")
    row_to_nct = load_excel_row_to_nct(EXCEL_PATH)
    annotator_map = load_annotator_workload(row_to_nct)
    print_annotator_analysis(agent_data, human_data, common_ncts, annotator_map)

    # Raw values for manual inspection
    print(f"\n{'=' * 115}")
    print("RAW VALUE COMPARISON (for manual inspection)")
    print(f"{'=' * 115}")
    for nct in common_ncts:
        print(f"\n{nct}:")
        for field_name in FIELDS:
            agent_raw = agent_data[nct].get(field_name, "")
            r1_raw = human_data[nct]["r1"].get(field_name)
            r2_raw = human_data[nct]["r2"].get(field_name)
            a_norm, _ = normalise(agent_raw, field_name)
            r1_norm, _ = normalise(r1_raw, field_name)
            r2_norm, _ = normalise(r2_raw, field_name)
            print(f"  {field_name}:")
            print(f"    Agent: {repr(agent_raw):>45} -> {repr(a_norm)}")
            print(f"    R1:    {repr(r1_raw):>45} -> {repr(r1_norm)}")
            print(f"    R2:    {repr(r2_raw):>45} -> {repr(r2_norm)}")

    # Save JSON results
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    json_output = {
        "timestamp": datetime.now().isoformat(),
        "source": "json_overnight_jobs",
        "excel_path": str(EXCEL_PATH),
        "job_files": [str(p) for p in json_paths],
        "common_ncts": common_ncts,
        "n_agent": len(agent_data),
        "n_human": len(human_data),
        "n_overlap": len(common_ncts),
        "fields": {},
        "disagreements": all_disagreements,
    }

    for field_name in FIELDS:
        field_json = {}
        for pair_name in ["Agent vs R1", "Agent vs R2", "R1 vs R2"]:
            pr = results[field_name][pair_name]
            kappa_ci = pr.get("kappa_ci")
            ac1_ci = pr.get("ac1_ci")
            field_json[pair_name] = {
                "n": pr["n"],
                "skipped_blank": pr["skipped_blank"],
                "agreements": pr["agreements"],
                "raw_agreement_pct": pr["raw_agreement_pct"],
                "cohens_kappa": pr["cohens_kappa"],
                "kappa_ci_lower": kappa_ci[0] if kappa_ci else None,
                "kappa_ci_upper": kappa_ci[1] if kappa_ci else None,
                "ac1": pr.get("ac1"),
                "ac1_ci_lower": ac1_ci[0] if ac1_ci else None,
                "ac1_ci_upper": ac1_ci[1] if ac1_ci else None,
                "prevalence_index": pr.get("prevalence_index"),
                "bias_index": pr.get("bias_index"),
                "n_disagreements": len(pr["disagreements"]),
            }

            # v3 three-tier results
            v3_pr = v3_results.get(field_name, {}).get(pair_name, {})
            field_json[pair_name]["tier1"] = v3_pr.get("tier1", {})
            field_json[pair_name]["tier2"] = v3_pr.get("tier2", {})
            field_json[pair_name]["tier3"] = v3_pr.get("tier3", {})
            field_json[pair_name]["v3_counts"] = v3_pr.get("counts", {})

        field_json["coverage"] = coverage_report.get(field_name, {})
        json_output["fields"][field_name] = field_json

    json_path = OUTPUT_DIR / "concordance_jobs_results.json"
    with open(json_path, "w") as f:
        json.dump(json_output, f, indent=2)
    print(f"\n\nJSON results saved to: {json_path}")


if __name__ == "__main__":
    main()
